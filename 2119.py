import pandas as pd
import re
from pathlib import Path
from openpyxl.utils import get_column_letter

# =========================
# 📂 RUTAS
# =========================

BASE = Path("base")
ARCHIVOS = Path("archivos")

ruta_catalogo = BASE / "catalogo.csv"
ruta_balance = ARCHIVOS / "balance_cierre.xlsx"
ruta_ajustes = ARCHIVOS / "anexo_ajustes.xlsx"
ruta_activos = ARCHIVOS / "resto_activos.xlsx"

ruta_reporte = BASE / "reporte_cambios.xlsx"

# =========================
# 🧹 LIMPIEZA
# =========================

def limpiar_columnas(df):
    df.columns = df.columns.str.lower().str.strip().str.replace(" ", "_")
    return df


def separar_cuenta(texto):

    if pd.isna(texto):
        return None, None, None

    texto = str(texto).replace(" ", "")

    match = re.match(r"([A-Za-z]+)(\d{1,3})(\d+)", texto)

    if match:
        return match.group(1), match.group(2), match.group(3)

    return None, None, None


# =========================
# 📥 CARGAR CSV BASE
# =========================

print("📄 Leyendo catálogo...")

df_catalogo = pd.read_csv(ruta_catalogo)
df_catalogo = limpiar_columnas(df_catalogo)

df_catalogo["fecha_de_creacion"] = pd.to_datetime(df_catalogo["fecha_de_creacion"])

ultima_fecha = df_catalogo["fecha_de_creacion"].max()
df_actual = df_catalogo[df_catalogo["fecha_de_creacion"] == ultima_fecha].copy()

print("Última versión:", ultima_fecha)


# =========================
# 📊 BALANCE
# =========================

def procesar_balance():

    print("📊 Procesando BALANCE...")

    archivo = pd.ExcelFile(ruta_balance)
    lista = []

    for hoja in archivo.sheet_names:

        df = pd.read_excel(archivo, sheet_name=hoja, header=1)
        df = limpiar_columnas(df)

        if "cuenta" in df.columns:

            df = df.reset_index()
            df.rename(columns={"index": "fila_excel"}, inplace=True)

            df["fila_excel"] += 2

            col_monto = df.columns.get_loc("monto")
            df["columna_excel"] = get_column_letter(col_monto + 1)

            df = df[["fila_excel","columna_excel","cuenta","descripcion","monto"]]

            df["archivo_origen"] = "balance"
            df["hoja_origen"] = hoja
            df["categoria_3"] = None

            lista.append(df)

    return pd.concat(lista, ignore_index=True)


# =========================
# 📊 AJUSTES
# =========================

def procesar_ajustes():

    print("📊 Procesando AJUSTES...")

    df = pd.read_excel(
        ruta_ajustes,
        sheet_name="detalle ajuste",
        header=4
    )

    df = limpiar_columnas(df)

    df = df.reset_index()
    df.rename(columns={"index": "fila_excel"}, inplace=True)

    df["fila_excel"] += 5

    df = df.rename(columns={"cuentas":"cuenta"})

    col_monto = df.columns.get_loc("monto")
    df["columna_excel"] = get_column_letter(col_monto + 1)

    df = df[["fila_excel","columna_excel","cuenta","descripcion","monto"]].dropna()

    df["archivo_origen"] = "ajustes"
    df["hoja_origen"] = "detalle ajuste"
    df["categoria_3"] = "Ajuste"

    return df


# =========================
# 📊 ACTIVOS
# =========================

def procesar_activos():

    print("📊 Procesando ACTIVOS...")

    hojas = ["resto de activos", "resto de contingentes"]
    lista = []

    for hoja in hojas:

        df = pd.read_excel(
            ruta_activos,
            sheet_name=hoja,
            header=None
        )

        df = df.reset_index()
        df.rename(columns={"index": "fila_excel"}, inplace=True)

        df["fila_excel"] += 3

        df = df.iloc[4:, [0,2,3,8]]
        df.columns = ["fila_excel","cuenta_raw","descripcion","monto"]

        df["columna_excel"] = "H"

        df["cuenta"] = df["cuenta_raw"].where(
            df["cuenta_raw"].astype(str).str.contains(r"\d")
        )

        df["cuenta"] = df["cuenta"].ffill()

        df = df[df["descripcion"].notna()]

        df[["categoria_1","categoria_2","cuenta"]] = df["cuenta"].apply(
            lambda x: pd.Series(separar_cuenta(x))
        )

        df["archivo_origen"] = "activos"
        df["hoja_origen"] = hoja
        df["categoria_3"] = "resto de activos"

        lista.append(df)

    return pd.concat(lista, ignore_index=True)


# =========================
# 🔄 ETL COMPLETO
# =========================

balance = procesar_balance()
ajustes = procesar_ajustes()
activos = procesar_activos()

df_nuevo = pd.concat([balance, ajustes, activos], ignore_index=True)

print("Total registros nuevos:", len(df_nuevo))


# =========================
# 🧩 COMPLETAR COLUMNAS
# =========================

for col in df_catalogo.columns:
    if col not in df_nuevo.columns:
        df_nuevo[col] = None


# =========================
# 🔍 COMPARACION
# =========================

merge = df_actual.merge(
    df_nuevo,
    on=["cuenta","categoria_3"],
    how="outer",
    suffixes=("_old","_new"),
    indicator=True
)

cambios = merge[
    (merge["_merge"] == "both") &
    (merge["monto_old"] != merge["monto_new"])
].copy()

print("Cambios detectados:", len(cambios))


# =========================
# 📍 CELDA EXACTA
# =========================

if not cambios.empty:

    cambios["celda_excel"] = (
        cambios["columna_excel"] + cambios["fila_excel"].astype(str)
    )

    reporte = cambios[[
        "cuenta",
        "categoria_3",
        "monto_old",
        "monto_new",
        "archivo_origen",
        "hoja_origen",
        "celda_excel"
    ]]

    reporte.to_excel(ruta_reporte, index=False)

    print("📄 Reporte generado:", ruta_reporte)


# =========================
# 🔄 ACTUALIZAR CATALOGO
# =========================

if not cambios.empty:

    print("🛠 Actualizando catálogo...")

    df_actualizado = df_actual.copy()

    for _, row in cambios.iterrows():

        filtro = (
            (df_actualizado["cuenta"] == row["cuenta"]) &
            (df_actualizado["categoria_3"] == row["categoria_3"])
        )

        df_actualizado.loc[filtro, "monto"] = row["monto_new"]

    df_actualizado["fecha_de_creacion"] = pd.Timestamp.today()

    df_final = pd.concat([df_catalogo, df_actualizado], ignore_index=True)

    df_final.to_csv(ruta_catalogo, index=False)

    print("✅ Nueva versión creada")

else:

    print("✅ SIN CAMBIOS - catálogo intacto")