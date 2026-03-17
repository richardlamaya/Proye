import pandas as pd
import re
from pathlib import Path
from openpyxl.utils import get_column_letter

# =========================
# 📂 RUTAS
# =========================

BASE = Path("base")
ARCHIVOS = Path("archivos")

ruta_catalogo = BASE / "catalogo.csv"
ruta_balance = ARCHIVOS / "balance_cierre.xlsx"
ruta_ajustes = ARCHIVOS / "anexo_ajustes.xlsx"
ruta_activos = ARCHIVOS / "resto_activos.xlsx"

ruta_reporte = BASE / "reporte_cambios.xlsx"

# =========================
# 🔎 DETECTAR HEADER
# =========================

def detectar_header(df):

    palabras = ["cuenta", "descripcion", "monto"]

    mejor_fila = None
    mejor_score = 0

    for i in range(min(15, len(df))):

        fila = df.iloc[i].astype(str).str.lower()

        score = sum(
            any(p in str(valor) for valor in fila)
            for p in palabras
        )

        if score > mejor_score:
            mejor_score = score
            mejor_fila = i

    if mejor_score >= 2:
        return mejor_fila

    return None


# =========================
# 🔎 DETECTAR COLUMNAS
# =========================

def detectar_columnas(df):

    columnas = df.columns

    col_cuenta = None
    col_desc = None
    col_monto = None

    for col in columnas:

        c = str(col).lower()

        if any(x in c for x in ["cuenta","codigo","cod"]):
            col_cuenta = col

        elif any(x in c for x in ["desc","nombre"]):
            col_desc = col

        elif any(x in c for x in ["monto","saldo","valor"]):
            col_monto = col

    return col_cuenta, col_desc, col_monto


# =========================
# 📥 LEER EXCEL LIMPIO
# =========================

def leer_excel_limpio(ruta, hoja):

    df_raw = pd.read_excel(ruta, sheet_name=hoja, header=None)

    header_row = detectar_header(df_raw)

    if header_row is None:
        raise ValueError(f"No se encontró header en {hoja}")

    df = df_raw.iloc[header_row+1:].copy()
    df.columns = df_raw.iloc[header_row]

    df.columns = (
        df.columns
        .astype(str)
        .str.lower()
        .str.strip()
        .str.replace(" ", "_")
    )

    df = df.loc[:, ~df.columns.str.contains("^unnamed")]

    return df, header_row


# =========================
# 🧩 PARSEAR CUENTA
# =========================

def separar_cuenta(texto):

    if pd.isna(texto):
        return None, None, None

    texto = str(texto).replace(" ", "")

    match = re.match(r"([A-Za-z]+)(\d{1,3})(\d+)", texto)

    if match:
        return match.group(1), match.group(2), match.group(3)

    return None, None, None


# =========================
# 📄 CSV BASE
# =========================

df_catalogo = pd.read_csv(ruta_catalogo)
df_catalogo.columns = df_catalogo.columns.str.lower().str.replace(" ", "_")

df_catalogo["fecha_de_creacion"] = pd.to_datetime(df_catalogo["fecha_de_creacion"])

ultima_fecha = df_catalogo["fecha_de_creacion"].max()
df_actual = df_catalogo[df_catalogo["fecha_de_creacion"] == ultima_fecha].copy()

# =========================
# 📊 BALANCE
# =========================

def procesar_balance():

    archivo = pd.ExcelFile(ruta_balance)
    lista = []

    for hoja in archivo.sheet_names:

        try:
            df, header_row = leer_excel_limpio(ruta_balance, hoja)
        except:
            continue

        col_cuenta, col_desc, col_monto = detectar_columnas(df)

        if not col_cuenta or not col_monto:
            continue

        df = df.reset_index()
        df.rename(columns={"index":"fila_excel"}, inplace=True)

        df["fila_excel"] += (header_row + 2)

        col_idx = df.columns.get_loc(col_monto)
        df["columna_excel"] = get_column_letter(col_idx + 1)

        df = df[[ "fila_excel","columna_excel", col_cuenta, col_desc, col_monto]]

        df.columns = ["fila_excel","columna_excel","cuenta","descripcion","monto"]

        df["archivo_origen"] = "balance"
        df["hoja_origen"] = hoja
        df["categoria_3"] = None

        lista.append(df)

    return pd.concat(lista, ignore_index=True)


# =========================
# 📊 AJUSTES
# =========================

def procesar_ajustes():

    df, header_row = leer_excel_limpio(ruta_ajustes, "detalle ajuste")

    col_cuenta, col_desc, col_monto = detectar_columnas(df)

    df = df.reset_index()
    df.rename(columns={"index":"fila_excel"}, inplace=True)

    df["fila_excel"] += (header_row + 2)

    col_idx = df.columns.get_loc(col_monto)
    df["columna_excel"] = get_column_letter(col_idx + 1)

    df = df[[ "fila_excel","columna_excel", col_cuenta, col_desc, col_monto]]

    df.columns = ["fila_excel","columna_excel","cuenta","descripcion","monto"]

    df["archivo_origen"] = "ajustes"
    df["hoja_origen"] = "detalle ajuste"
    df["categoria_3"] = "Ajuste"

    return df


# =========================
# 📊 ACTIVOS
# =========================

def procesar_activos():

    hojas = ["resto de activos", "resto de contingentes"]
    lista = []

    for hoja in hojas:

        df_raw = pd.read_excel(ruta_activos, sheet_name=hoja, header=None)

        df = df_raw.iloc[4:, [2,3,8]]
        df.columns = ["cuenta_raw","descripcion","monto"]

        df = df.reset_index()
        df.rename(columns={"index":"fila_excel"}, inplace=True)

        df["fila_excel"] += 5
        df["columna_excel"] = "H"

        df["cuenta"] = df["cuenta_raw"].ffill()

        df[["categoria_1","categoria_2","cuenta"]] = df["cuenta"].apply(
            lambda x: pd.Series(separar_cuenta(x))
        )

        df["archivo_origen"] = "activos"
        df["hoja_origen"] = hoja
        df["categoria_3"] = "resto de activos"

        lista.append(df[["fila_excel","columna_excel","cuenta","descripcion","monto","categoria_1","categoria_2","categoria_3","archivo_origen","hoja_origen"]])

    return pd.concat(lista, ignore_index=True)


# =========================
# 🔄 ETL
# =========================

balance = procesar_balance()
ajustes = procesar_ajustes()
activos = procesar_activos()

df_nuevo = pd.concat([balance, ajustes, activos], ignore_index=True)

# =========================
# 🔍 COMPARACION
# =========================

merge = df_actual.merge(
    df_nuevo,
    on=["cuenta","categoria_3"],
    how="outer",
    suffixes=("_old","_new"),
    indicator=True
)

cambios = merge[
    (merge["_merge"] == "both") &
    (merge["monto_old"] != merge["monto_new"])
].copy()

# =========================
# 📍 CELDA
# =========================

cambios["celda_excel"] = cambios["columna_excel"] + cambios["fila_excel"].astype(str)

reporte = cambios[[
    "cuenta","categoria_3","monto_old","monto_new",
    "archivo_origen","hoja_origen","celda_excel"
]]

reporte.to_excel(ruta_reporte, index=False)

print("Reporte generado")